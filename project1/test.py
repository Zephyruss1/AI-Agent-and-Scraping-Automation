from openpyxl import load_workbook
import os
import urllib.request
import concurrent.futures
from playwright.async_api import async_playwright
import os
import time
from typing import Optional
import asyncio

class DownloadPDF:
    def __init__(self, excel_path):
        self.excel_path = excel_path

    def _load_excel(self):
        wb = load_workbook(self.excel_path)
        ws = wb.active
        if ws:
            return ws


    def download_arxiv_pdf(self, pdf_url: str, save_dir: str ="pdfs"):
        if not pdf_url:
            print(f"⚠️ [WARNING] Skipping invalid paper url: {pdf_url}")
            return False

        if "arxiv.org/abs" in pdf_url:
            pdf_url = pdf_url.replace("arxiv.org/abs", "arxiv.org/pdf") + ".pdf"

        save_path = os.path.join(save_dir, f"{pdf_url.split('/')[-1]}")

        os.makedirs(save_dir, exist_ok=True)

        try:
            req = urllib.request.Request(pdf_url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req) as response:
                content_type = response.headers.get("Content-Type")

                if content_type != "application/pdf":
                    print(f"❌ [ERROR] {pdf_url}: PDF can't be downloaded. Content-Type: {content_type}")
                    return False

                if pdf_url.split("/")[-1] in os.listdir(save_dir):
                    print(f"⚠️ [WARNING] {pdf_url}: PDF already exists in the directory.")
                    return False

                with open(save_path, "wb") as f:
                    f.write(response.read())

            print(f"    ✅ [INFO] {pdf_url}: PDF successfully downloaded -> {save_path}")
            return True

        except Exception as e:
            print(f"❌ [ERROR] {pdf_url}: {e}")
            return False

    def start_download(self, max_workers: int = 3, max_download_num: int = None):
        print("\n📍 Step 6: Downloading PDFs!")
        ws = self._load_excel()
        if not ws:
            print("❌ [ERROR] Excel file not found!")
            return
        pdf_urls = [row[3].value for row in ws.iter_rows(min_row=2, max_col=4)]

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(self.download_arxiv_pdf, pdf_url): pdf_url for pdf_url in pdf_urls}

            for future in concurrent.futures.as_completed(futures):
                pdf_url = futures[future]
                try:
                    future.result()
                except Exception as e:
                    print(f"❌ [ERROR] {pdf_url}: {e}")


def sync_main():
    pdf_downloader = DownloadPDF("arxiv_scraped_data.xlsx")
    pdf_downloader.start_download()

if __name__ == "__main__":
    sync_main()

class ChatGPTPlaywright:
    def __init__(self, profile_path: Optional[str] = None):
        self.profile_path = profile_path

    async def start(self):
        self.playwright = await async_playwright().start()
        self.browser = await self._launch_browser()
        self.page = await self.browser.new_page()
        await self.page.goto("https://www.perplexity.ai/")
        time.sleep(5)
        await self._interact_with_interface()
        await self.page.screenshot(path="perplexity.png")
        time.sleep(10)

    async def _launch_browser(self):
        proxy = {
            "server": "brd.superproxy.io:33335",
            "username": "brd-customer-hl_fc29e1f2-zone-semanticscholar",
            "password": "p8tgqocdlqav"
        }
        # Configure browser launch options
        browser_args = [
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--profile-directory=Profile 1",
        ]

        # Add user profile directory if provided
        if self.profile_path:
            abs_profile_path = os.path.expanduser(self.profile_path)
            print(f"ℹ️ [INFO] Using profile directory: {abs_profile_path}")

        browser = await self.playwright.chromium.launch_persistent_context(
            user_data_dir=abs_profile_path if self.profile_path else None,
            headless=False,
            args=browser_args,
            proxy=proxy,
        )
        return browser

    async def _interact_with_interface(self):
        await self.page.click('xpath=//*[@id="__next"]/main/div/div/div[1]/div/div/div/div[4]/div/button[2]')
        time.sleep(5)
        await self.page.click('xpath=//*[@id="__next"]/div[2]/div[2]/div/div/div/div[2]/div[2]/div[2]'
                              '/div/div[1]/button[1]')
        time.sleep(5)

    async def close(self):
        """Close the browser and stop Playwright."""
        if hasattr(self, "browser") and self.browser:
            await self.browser.close()
        if hasattr(self, "playwright") and self.playwright:
            await self.playwright.stop()


async def main():
    profile_path = r"C:\clonedUserData"

    chat_gpt = ChatGPTPlaywright(profile_path=profile_path)
    await chat_gpt.start()
    try:
        print("ChatGPT page loaded successfully!")
    finally:
        await chat_gpt.close()


# if __name__ == "__main__":
#     asyncio.run(main())