from typing import List, Dict, Optional
from openpyxl import Workbook, load_workbook
import asyncio
from playwright.async_api import async_playwright
import aiohttp


async def _load_excel():
    wb = load_workbook(filename='arxiv_scraped_data.xlsx')
    ws = wb.active
    return ws


def _save_excel(author_list: List[Dict], keyword: List[str]) -> None:
    if not author_list:
        print("     [INFO] No author details found!")
        return

    wb = Workbook()
    ws = wb.active
    ws.append(["Author", "AuthorID", "PaperID", "LastPaperLink", "AmountOfMentions", "Keyword"])

    for author in author_list:
        ws.append([
            author.get("Author", "null"),
            author.get("AuthorID", "null"),
            author.get("PaperID", "null"),
            author.get("LastPaperLink", "null"),
            author.get("AmountOfMentions", "null"),
            ", ".join(author.get("Keyword", ["null"])) if isinstance(author.get("Keyword", list), list) else author.get(
                "Keyword", "null")
        ])

    try:
        wb.save("arxiv_scraped_data.xlsx")
        print("    ✅ [INFO] Details successfully saved to Excel!")
    except Exception as e:
        print(f"    ❌ [ERROR] An error occurred while saving Excel file: {e}")


class ArxivScraper:
    def __init__(self, keyword: Optional[List[str]] = None,
                 OR: Optional[List[str]]  = None,
                 NOT: Optional[List[str]] = None,
                 date_from: str = None,
                 date_to: str = None):
        self.page = None
        self.keyword = keyword if keyword is not None else []
        self.OR = OR if OR is not None else []
        self.NOT = NOT if NOT is not None else []
        self.date_from = date_from
        self.date_to = date_to
        self.papers: List[Dict] = []

    async def pagination(self) -> bool:
        print("     [INFO] Checking for next page...")
        next_page_locator = self.page.locator('xpath=//*[@id="main-container"]/div[2]/nav[1]/a[2]')
        try:
            await next_page_locator.wait_for(state='visible', timeout=10000)
            next_class = await next_page_locator.get_attribute("class")
            if next_class and "disabled" in next_class:
                print("    ✅ [INFO] No more pages.")
                return False
            await next_page_locator.click()
            await asyncio.sleep(1)
            return True
        except Exception as e:
            print(f"     [INFO] End of the page.")
            return False

    async def connect_to_arxiv_and_search(self):
        print("\n📍 Step 1: Connecting to Arxiv!")

        url = "https://arxiv.org/search/advanced"

        try:
            await self.page.goto(url, wait_until='load')
            await asyncio.sleep(2)
            print("     ✅ [INFO] Arxiv connection successful!")
        except Exception as e:
            print(f"     ❌ [ERROR] An error occurred during navigation: {e}")

        try:
            if self.date_from or self.date_to:
                date_range = self.page.locator('xpath=//*[@id="date-filter_by-3"]')
                await date_range.click()
                date_from_box = self.page.locator('xpath=//*[@id="date-from_date"]')
                date_to_box = self.page.locator('xpath=//*[@id="date-to_date"]')
                await date_from_box.fill(self.date_from)
                await asyncio.sleep(1)
                await date_to_box.fill(self.date_to)
                await asyncio.sleep(1)

            # Combine all search terms (keyword, OR, NOT)
            search_terms = []
            if self.keyword:
                search_terms.extend(self.keyword if isinstance(self.keyword, list) else [self.keyword])
            if self.OR:
                search_terms.extend(self.OR if isinstance(self.OR, list) else [self.OR])
            if self.NOT:
                search_terms.extend(self.NOT if isinstance(self.NOT, list) else [self.NOT])

            # Fill the first search box (already present)
            search_box = self.page.locator(f'xpath=//*[@id="terms-0-term"]')
            await search_box.fill(search_terms[0])

            # Handle additional search terms dynamically
            for i, term in enumerate(search_terms[1:], start=1):
                add_button_xpath = f'//*[@id="terms-fieldset"]/fieldset/div[{i + 1}]/div/button[1]'
                another_term_button = self.page.locator(f'xpath={add_button_xpath}')

                await another_term_button.click()
                await asyncio.sleep(1)

                dropdown_xpath = self.page.locator(f'xpath=//*[@id="terms-{i}-operator"]')
                if term in self.OR:
                    await dropdown_xpath.select_option("OR")
                elif term in self.NOT:
                    await dropdown_xpath.select_option("NOT")
                else:
                    await dropdown_xpath.select_option("AND")

                new_search_box = self.page.locator(f'xpath=//*[@id="terms-{i}-term"]')
                await new_search_box.fill(term)
                print(f"     [INFO] Entered: {term}")

            await search_box.press('Enter')
        except Exception as e:
            print(f"     ❌ [ERROR] An error occurred while searching: {e}")

    async def get_paper_ids(self, max_pages_DEBUG_MODE: Optional[int] = None) -> List[Dict]:
        print("\n📍 Step 2: Scraping paper links!")
        page_count = 0
        if max_pages_DEBUG_MODE:
            print(f"     ⚠️ [INFO] Debug mode enabled. Scraping only {max_pages_DEBUG_MODE} pages.")

        while True:
            page_count += 1
            if max_pages_DEBUG_MODE and page_count > max_pages_DEBUG_MODE:
                break

            try:
                paper_entries = self.page.locator('xpath=//*[@id="main-container"]/div[2]/ol/li')
            except Exception as e:
                print(f"     ❌ [ERROR] Waiting for paper entries failed: {e}")
                break

            papers = await paper_entries.all()
            if not papers:
                print("     [INFO] No paper entries found on this page.")
                break

            for paper in papers:
                try:
                    paper_link = await (
                        paper.locator('xpath=//p[@class="list-title is-inline-block"]/a').get_attribute("href"))
                    author_locator = paper.locator('xpath=//p[contains(@class, "authors")]')
                    author_text = await author_locator.text_content()

                    paper_id = None
                    if paper_link:
                        paper_id = paper_link.split("/")[-1]

                    if not paper_id:
                        print(f"     ⚠️ [WARNING] Paper ID missing for one entry, skipping...")
                        continue

                    if author_text:
                        author_names = author_text.split(":")[1].split(",")
                        for author in author_names:
                            self.papers.append({
                                "PaperID": paper_id,
                                "Author": author.strip(),
                                "AuthorID": 0,
                                "LastPaperLink": "",
                                "AmountOfMentions": 0,
                                "Keyword": self.keyword
                            })

                except Exception as e:
                    print(f"     ❌ [ERROR] Failed to process a paper entry: {e}")

            print(f"     [INFO] Scraped page {page_count}")
            if not await self.pagination():
                break

        print("     ✅ [INFO] Paper links scraped successfully!")
        return self.papers

    async def get_author_details(self) -> Optional[List[Dict]]:
        print("\n📍 Step 3: Scraping author details!")
        if not self.papers:
            print("     [INFO] No paper links found!")
            return None
        async with aiohttp.ClientSession() as session:
            for paper in self.papers:
                await asyncio.sleep(2)
                api = (f"https://api.semanticscholar.org/v1/paper/arXiv:{paper['PaperID']}"
                       f"?include_unknown_references=true")
                try:
                    async with session.get(api) as response:
                        response.raise_for_status()
                        data = await response.json()

                        if 'authors' not in data:
                            print(f"    [WARNING] No authors found for paper {paper['PaperID']}")
                            continue

                        for author in data['authors']:
                            if author.get("name", "").split(" ")[-1] == paper["Author"].split(" ")[-1]:
                                paper["AuthorID"] = author.get("authorId", "null")

                        print(f"     [INFO] Processed authors for paper {paper['PaperID']}")
                except aiohttp.ClientResponseError as e:
                    print(f"    ❌ [ERROR] HTTP error for paper {paper['PaperID']}: {e}")
                except KeyError as e:
                    print(f"    ❌ [ERROR] Missing key in response for paper {paper['PaperID']}: {e}")
                except Exception as e:
                    print(f"    ❌ [ERROR] An error occurred for paper {paper['PaperID']}: {e}")
                await asyncio.sleep(1)

        if not any(paper["AuthorID"] for paper in self.papers):
            print("     [INFO] No author details found for any papers.")
            return None

        print("    ✅ [INFO] Author details scraped successfully!")
        return self.papers

    async def get_amount_of_mentions(self) -> Optional[List[Dict]]:
        print("\n📍 Step 4: Using Scholar API to extract amount of mentions!")
        if not self.papers:
            print("     [INFO] No author details found!")
            return None

        async with aiohttp.ClientSession() as session:
            for row_info in self.papers:
                author_id = row_info.get("AuthorID")
                if author_id == "null":
                    print(f"    [INFO] No valid author name for {row_info['Author']}. Skipping.")
                    continue

                # Create the API URL with the first author's authorId (or pick another strategy to query)
                api = (f"https://api.semanticscholar.org/graph/v1/author/{author_id}"
                       f"/papers?fields=url,publicationDate,authors&limit=1000")
                try:
                    async with session.get(api) as response:
                        response.raise_for_status()
                        data = await response.json()

                        if 'data' in data and data['data']:
                            amount_of_mentions = data['data']  # List of papers
                            print(f"    📄 {row_info['Author']} -> {len(amount_of_mentions)} papers found.")

                            # Update the amount of mentions for the author
                            row_info["AmountOfMentions"] = len(amount_of_mentions)
                        else:
                            print(f"    ⚠️ [WARNING] No papers found for author {row_info['Author']}")

                        await asyncio.sleep(2)  # Sleep to prevent hitting rate limits

                except aiohttp.ClientResponseError as e:
                    print(f"    ❌ [ERROR] HTTP error for author {row_info['Author']}: {e}")
                except KeyError as e:
                    print(f"    ❌ [ERROR] Missing key in response for author {row_info['Author']}: {e}")
                except Exception as e:
                    print(f"    ❌ [ERROR] An error occurred for author {row_info['Author']}: {e}")

        print("    ✅ [INFO] Related papers scraped successfully!")
        return self.papers

    async def get_related_paper_details(self):
        print("\n📍 Step 5: Scraping related paper details!")
        if not self.papers:
            print("     [INFO] No author details found!")
            return None

        for row in self.papers:
            author = row.get("Author")
            if not author:
                print(f"     [INFO] No author name found for row: {row}")
                continue

            try:
                # Enter the author's name in the search box
                search_box = self.page.locator('xpath=/html/body/header/div[2]/div[2]/form/div/div[1]/input')
                await search_box.fill(author)
                await search_box.press('Enter')
                await asyncio.sleep(2)  # Wait for the search results to load

                # Locate the latest paper
                paper_entries = self.page.locator('xpath=//*[@id="main-container"]/div[2]/ol/li')
                papers = await paper_entries.all()
                if not papers:
                    print(f"     [INFO] No papers found for author: {author}")
                    continue

                # Get the link of the first paper (latest paper)
                try:
                    paper_link = await papers[0].locator(
                        'xpath=//p[@class="list-title is-inline-block"]/a'
                    ).get_attribute("href")
                    if paper_link:
                        print(f"     🔗 Found paper link for {author}: {paper_link}")
                        row["LastPaperLink"] = paper_link
                    else:
                        print(f"     ⚠️ [WARNING] No paper link found for {author}")
                except Exception as e:
                    print(f"     ❌ [ERROR] Failed to process the first paper entry for {author}: {e}")

            except Exception as e:
                print(f"     ❌ [ERROR] An error occurred while processing {author}: {e}")

        print("    ✅ [INFO] Last paper links scraped successfully!")
        return self.papers


async def async_main():
    proxy = {
        "server": "brd.superproxy.io:33335",
        "username": "brd-customer-hl_fc29e1f2-zone-isp_proxy1",
        "password": "53b6d13ej50m"
    }
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1280, 'height': 720})
        page = await context.new_page()

        # Create the ArxivScraper instance and assign the open page
        arxiv_scraper = ArxivScraper(keyword=["photonic circuits"],
                                     date_from="2022-01-01", date_to="2023-12-31")
        arxiv_scraper.page = page

        # Step 1: Connect and search
        await arxiv_scraper.connect_to_arxiv_and_search()

        # Step 2: Get paper IDs (use max_pages_DEBUG_MODE=1 for testing)
        await arxiv_scraper.get_paper_ids() #max_pages_DEBUG_MODE=1

        # Step 3: Get author details
        await arxiv_scraper.get_author_details()

        # Step 4: Get amount of mentions
        await arxiv_scraper.get_amount_of_mentions()

        # Step 5: Get related paper details
        await arxiv_scraper.get_related_paper_details()

        # Save the scraped details to an Excel file
        _save_excel(arxiv_scraper.papers, arxiv_scraper.keyword)


if __name__ == "__main__":
    asyncio.run(async_main())

# TODO: get_author_details request optimization
# TODO: Scraping author email if exists in paper details
# TODO: Scraping related tags from paper details to filter author last paper

#job title, email, organization