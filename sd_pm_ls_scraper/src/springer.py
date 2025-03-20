import asyncio
import os
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright


async def _save_excel_all_springer(articles: List[Dict]) -> None:
    if not articles:
        print("     [INFO] No articles to save!")
        return

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "title",
            "link",
            "description",
            "authors",
            "published_date",
            "keyword",
        ]
    )
    # "title": title.strip() if title else "No title found",
    # "link": f"{self.url}{title_link}",
    # "description": description.strip() if description else "No description found",
    # "authors": authors.strip() if authors else "No authors found",
    # "published_date": published_date.strip() if published_date else "No published date found"
    for article in articles:
        ws.append(
            [
                article.get("title", "null"),
                article.get("link", "null"),
                article.get("description", "null"),
                article.get("authors", "null"),
                article.get("published_date", "null"),
                article.get("keyword", "null"),
            ]
        )

    try:
        save_path = os.path.abspath("springer_results.xlsx")
        wb.save(save_path)
        print(
            f"    ✅ [INFO] {len(articles)} articles successfully saved to Excel at {save_path}!"
        )
    except Exception as e:
        print(f"    ❌ [ERROR] An error occurred while saving Excel file: {e}")


async def _load_excel(file_name: str, sheet_name=None, page_index=0):
    """
    Load an Excel workbook and return the specified worksheet.

    Args:
        sheet_name (str, optional): Name of the sheet to load. If None, uses page_index.
        page_index (int, optional): Index of the sheet to load (0-based). Default is 0.

    Returns:
        Worksheet: The specified worksheet from the Excel file.
    """
    wb = load_workbook(filename=file_name)

    if sheet_name and sheet_name in wb.sheetnames:
        # Access by sheet name if provided and valid
        ws = wb[sheet_name]
    else:
        # Access by index (default to first sheet if index is out of range)
        if page_index < len(wb.sheetnames):
            ws = wb.worksheets[page_index]
        else:
            # Default to the active sheet if specified page doesn't exist
            print(f"Warning: Sheet index {page_index} not found. Using active sheet.")
            ws = wb.active

    return ws


class SpringerScraper:
    def __init__(self, keyword: str):
        self.keyword = keyword
        self.url = "https://link.springer.com/"
        self.articles = []
        self.page = None

    async def _check_cookies(self, page):
        """Handle cookie consent dialog if it appears"""
        try:
            await page.wait_for_selector("xpath=/html/body/dialog", timeout=5000)
            close_button = await page.query_selector(
                "xpath=/html/body/dialog/div/div/div[3]/button"
            )
            if close_button:
                await close_button.click()
        except Exception as e:
            print(f"Cookie dialog not found or already closed: {e}")

    async def searching_and_gathering_papers(self, page):
        """Perform a search on Springer and gather the articles"""
        print("\n📍 Step 1: Connecting to Springer!")  # Moved inside the method

        await page.goto(self.url)
        await self._check_cookies(page)
        await page.wait_for_selector('xpath=//*[@id="homepage-search"]', timeout=10000)
        await page.fill('xpath=//*[@id="homepage-search"]', self.keyword)

        search_button = await page.query_selector(
            'xpath=//*[@id="main"]/div[1]/div/div/div[2]/search/form/div/button'
        )
        if search_button:
            await search_button.click()

        await page.wait_for_load_state("networkidle")

        await page.wait_for_selector(
            'xpath=//*[@id="main"]/div/div/div/div[2]/div[2]/ol', timeout=10000
        )
        articles = await page.query_selector_all(
            'xpath=//*[@id="main"]/div/div/div/div[2]/div[2]/ol/li'
        )
        if not articles:
            print("No articles found. Check if the selector is correct.")
            return
        print(f"Found {len(articles)} articles")

        if len(articles) == 0:
            print("No articles found. Check if the selector is correct.")
            return  # Added explicit return to prevent processing when no articles found

        print("\n📍 Step 2: Gathering articles!")
        for i, article in enumerate(articles[:5]):
            # Use explicit xpath selectors
            title_element = await article.query_selector("xpath=.//div/h3")
            title = (
                await title_element.text_content()
                if title_element
                else "No title found"
            )

            # For the link, find the anchor inside h3
            title_link_element = await article.query_selector("xpath=.//div/h3/a")
            title_link = (
                await title_link_element.get_attribute("href")
                if title_link_element
                else None
            )

            description_element = await article.query_selector("xpath=.//div/p")
            description = (
                await description_element.text_content()
                if description_element
                else "No description found"
            )

            authors_element = await article.query_selector("xpath=.//div/div[3]/div")
            authors = (
                await authors_element.text_content()
                if authors_element
                else "No authors found"
            )

            published_date_element = await article.query_selector(
                "xpath=.//div/div[4]/div/span[1]"
            )
            # published_date_element = await article.query_selector('xpath=//*[@id="main"]/div/div/div/div[2]/div[2]/ol/li[1]/div[1]/div[3]/div/span')
            published_date = (
                await published_date_element.text_content()
                if published_date_element
                else "No published date found"
            )

            # --- Print the articles --- #
            print(f"\nArticle {i + 1}:")
            print(f"Title: {title.strip() if title else 'No title found'}")
            print(f"Title Link: {self.url}{title_link}")
            print(
                f"Description: {description.strip() if description else 'No description found'}"
            )
            print(f"Authors: {authors.strip() if authors else 'No authors found'}")
            print(
                f"Published Date: {published_date.strip() if published_date else 'No published date found'}"
            )
            print("\n")

            # --- Append the articles to the list --- #
            self.articles.append(
                {
                    "title": title.strip() if title else "No title found",
                    "link": f"{self.url}{title_link}",
                    "description": description.strip()
                    if description
                    else "No description found",
                    "authors": authors.strip() if authors else "No authors found",
                    "published_date": published_date.strip()
                    if published_date
                    else "No published date found",
                }
            )

    async def similarity_finder(self, page):
        import re

        ws_page2 = await _load_excel(file_name="test_springer.xlsx", page_index=1)

        pattern = re.compile(r"\b" + re.escape(self.keyword) + r"\b", re.IGNORECASE)
        for _row_idx, link in enumerate(
            ws_page2.iter_rows(min_row=2, max_row=5, values_only=True), 2
        ):
            pdf_url = link[1]
            await page.goto(pdf_url)
        text = await page.content()  # Get the page content
        if pattern.search(text):
            print("Match found!")
        pass


async def async_springer():
    async with async_playwright() as pw:
        args = [
            "--no-sandbox",
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--disable-background-timer-throttling",
            "--disable-popup-blocking",
            "--disable-backgrounding-occluded-windows",
            "--disable-renderer-backgrounding",
            "--disable-window-activation",
            "--disable-focus-on-load",
            "--no-first-run",
            "--no-default-browser-check",
            "--no-startup-window",
            "--window-position=960,700",
        ]

        browser = await pw.chromium.launch(headless=False, args=args)
        context = await browser.new_context(viewport={"width": 1280, "height": 720})
        page = await context.new_page()

        all_articles = []

        try:
            ws_page2 = await _load_excel(
                file_name="sciencedirect_pubmed.xlsx", page_index=1
            )

            for _row_idx, author in enumerate(
                ws_page2.iter_rows(min_row=2, max_row=5, values_only=True), 2
            ):
                keyword = author[0]
                if not keyword:
                    print(f"Empty keyword at row {_row_idx}, skipping...")
                    continue

                print(f"\n{'=' * 50}")
                print(f"Processing keyword: {keyword} (Row {_row_idx})")
                print(f"{'=' * 50}")

                try:
                    scraper = SpringerScraper(keyword=keyword)
                    await scraper.searching_and_gathering_papers(page)

                    if scraper.articles:
                        print(
                            f"Found {len(scraper.articles)} articles for '{keyword}', adding to collection..."
                        )
                        # Add keyword to each article
                        for article in scraper.articles:
                            article["keyword"] = keyword
                        all_articles.extend(scraper.articles)
                    else:
                        print(
                            f"No articles found for keyword '{keyword}', moving to next keyword..."
                        )
                except Exception as e:
                    print(f"Error processing keyword '{keyword}': {e}")
                    os.makedirs("errors", exist_ok=True)
                    await page.screenshot(
                        path=f"errors/error_{keyword.replace(' ', '_')}.png"
                    )
                    print("Error screenshot saved. Continuing with next keyword...")

                await asyncio.sleep(2)

            if all_articles:
                print(f"\nSaving {len(all_articles)} total articles to Excel...")
                await _save_excel_all_springer(all_articles)
                print("No articles found for any keywords.")

        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            await page.screenshot(path="unexpected_error.png")

        finally:
            print("\nClosing browser...")
            await browser.close()


if __name__ == "__main__":
    import asyncio

    asyncio.run(async_springer())
