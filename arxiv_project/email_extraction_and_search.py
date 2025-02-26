import concurrent.futures
import os
import subprocess
from dataclasses import dataclass, field
from typing import List, Optional

import pdfplumber
import requests
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from pydantic import BaseModel
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

load_dotenv()


def _load_excel(file_name: str) -> object:
    """Load the excel file."""
    wb = load_workbook(filename=file_name)
    ws = wb.active
    return wb, ws


class DownloadPDF:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.proxy = {
            "account_id": "hl_fc29e1f2",
            "zone_name": "semanticscholar",
            "password": "p8tgqocdlqav",
        }

    def download_arxiv_pdf(self, pdf_url: str, save_dir: str = "pdfs") -> bool:
        """Download PDF from the given URL using requests."""
        if not pdf_url:
            print(f"    ⚠️ [WARNING] Skipping invalid paper url: {pdf_url}")
            return False

        # Convert abstract URL to PDF URL
        if "arxiv.org/abs" in pdf_url:
            pdf_url = pdf_url.replace("arxiv.org/abs", "arxiv.org/pdf") + ".pdf"

        save_path = os.path.join(save_dir, f"{pdf_url.split('/')[-1]}")
        os.makedirs(save_dir, exist_ok=True)

        if os.path.exists(save_path):
            print(f"    ⚠️ [WARNING] {pdf_url}: PDF already exists.")
            return True

        try:
            proxy_url = "brd.superproxy.io:33335"
            proxy_auth = f"brd-customer-{self.proxy['account_id']}-zone-{self.proxy['zone_name']}:{self.proxy['password']}"

            cmd = [
                "curl",
                "--proxy",
                proxy_url,
                "--proxy-user",
                proxy_auth,
                "-k",
                pdf_url,
                "-o",
                save_path,
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)
            if result.returncode != 0:
                print(
                    f"    ❌ [ERROR] {pdf_url}: PDF can't be downloaded. Error: {result.stderr}"
                )
                return False

            print(
                f"    ✅ [INFO] {pdf_url}: PDF successfully downloaded -> {save_path}"
            )
            return True

        except requests.exceptions.RequestException as e:
            print(f"    ❌ [ERROR] {pdf_url}: Request error - {str(e)}")
            return False

    def start_download(self, max_workers: int = 3) -> None:
        """Start downloading PDFs with rate limiting consideration."""
        print("\n📍 Step 6: Downloading PDFs!")
        wb, ws = _load_excel(file_name=self.excel_path)

        if not ws:
            print("     ❌ [ERROR] Excel file not found!")
            return

        pdf_urls = [row[3].value for row in ws.iter_rows(min_row=2, max_col=4)]
        print(f"    ℹ️ [INFO] Found {len(pdf_urls)} URLs to download.")

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(self.download_arxiv_pdf, pdf_url): pdf_url
                for pdf_url in pdf_urls
            }
            for future in concurrent.futures.as_completed(futures):
                pdf_url = futures[future]
                try:
                    future.result()
                except Exception as e:
                    print(f"    ❌ [ERROR] {pdf_url}: Thread error - {e}")


class LoadPDF:
    """Load PDFs and extract text from them."""

    print("\n📍 Step 7: Loading PDFs and Extracting Text!")

    def __init__(self):
        self.pdf_dir = "pdfs/"

    def extract_text_from_pdf(self) -> str:
        """Extract text from a PDF file."""
        list_pdf_files = self.list_pdf_files()
        print("    ✅ [INFO] Listing PDF files in the directory.")
        for pdf_file in list_pdf_files:
            text = ""
            with pdfplumber.open(f"{self.pdf_dir}{pdf_file}") as pdf:
                print(f"     [INFO] Extracting text from {pdf_file}")
                for page in pdf.pages:
                    text += page.extract_text() + "\n"
            return text

    def list_pdf_files(self) -> List[str]:
        """List all PDF files in the directory."""
        return [file for file in os.listdir(self.pdf_dir) if file.endswith(".pdf")]

    def __len__(self) -> int:
        """Return the number of PDF files."""
        return len(self.list_pdf_files())


@dataclass(frozen=True)
class ChatGPTConfig:
    """ChatGPT API Config Class."""

    model: str = "gpt-4o"
    api_key: str = field(default_factory=lambda: os.getenv("OPENAI_API_KEY", ""))
    temperature: float = 0.2


class ExtractEmails:
    """Implement ChatGPTConfig to ExtractEmails."""

    def __init__(self, config: Optional[ChatGPTConfig] = None) -> None:
        if config is None:
            config = ChatGPTConfig()
        self.config = config

    def chatgpt_response(self, pdf_texts: str) -> List[str]:
        """Extract email addresses from the text using ChatGPT."""
        print("\n📍 Step 8: [ChatGPT] Extracting Email Addresses!")
        client = OpenAI()
        client.api_key = self.config.api_key
        completion = client.chat.completions.create(
            model=self.config.model,
            store=True,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a data extraction assistant. The user will provide the text from a PDF. "
                        "Identify and list all email addresses that appear in the text (i.e., strings containing '@'). "
                        "If you find no email addresses, return 'None'. "
                        "Output only the emails or 'None'—no additional explanations."
                    ),
                },
                {"role": "user", "content": pdf_texts},
            ],
            temperature=self.config.temperature,
        )

        list_of_emails = completion.choices[0].message.content.split("\n")
        if list_of_emails:
            print(f"    ✅ [INFO] Email addresses listed: {list_of_emails}")
            return list_of_emails


@dataclass(frozen=True)
class PerplexityConfig:
    """Perplexity API Config Class."""

    url: str = "https://api.perplexity.ai/chat/completions"
    model: str = "sonar-pro"
    api_key: str = field(default_factory=lambda: os.getenv("PERPLEXITY_API_KEY", ""))

    def get_headers(self) -> dict:
        """Dynamically returning headers."""
        if not self.api_key:
            raise ValueError("PERPLEXITY_API_KEY is not set.")
        return {"Authorization": f"Bearer {self.api_key}"}


class AnswerFormat(BaseModel):
    """Answer Format for Perplexity API."""

    email_adress: str


class WebSearch:
    def __init__(self, name: str, config: Optional[PerplexityConfig] = None) -> None:
        if config is None:
            config = PerplexityConfig()
        """Implement PerplexityConfig to WebSearch."""
        self.config = config
        self.author_name = name
        wb, ws = _load_excel("arxiv_scraped_data_backup.xlsx")
        self.headers = [
            str(cell.value).strip().lower() if cell.value else None for cell in ws[1]
        ]

        self.keyword_index = None
        for i, header in enumerate(self.headers):
            if header == "keyword":
                self.keyword_index = i
                break

    def perplexity_search(self) -> str:
        """Search for email addresses for the provided author name."""
        print("\n📍 Step 8: [Perplexity] Extracting Email Addresses!")
        payload = {
            "model": self.config.model,
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "You are a web searcher assistant. The user will provide an author name. "
                        f"Your task is: Search email addresses in the {self.keyword_index} field for provided author name."
                        "If you find no email addresses, return 'None'. "
                        "Output only the emails or 'None'—no additional explanations."
                    ),
                },
                {"role": "user", "content": f"{self.author_name} email adress"},
            ],
            "response_format": {
                "type": "json_schema",
                "json_schema": {"schema": AnswerFormat.model_json_schema()},
            },
        }

        response = requests.post(
            self.config.url, headers=self.config.get_headers(), json=payload
        )

        if response.status_code == 200:
            try:
                response_json = response.json()
                print(response_json["choices"][0]["message"]["content"])
                list_of_emails = response_json["choices"][0]["message"][
                    "content"
                ].split("\n")
                if list_of_emails:
                    print(f"    ✅ [INFO] Email addresses listed: {list_of_emails}")
                    return list_of_emails

            except requests.JSONDecodeError:
                print("Error: Response content is not valid JSON")
        else:
            print(f"Error: Received status code {response.status_code}")

    def browser_use(self) -> str:
        """Search for email addresses using the browser-use."""
        print("\n📍 Step 8: [Browser-Use] Extracting Email Addresses!")

        try:
            import asyncio

            from browser_use import Agent, Browser, BrowserConfig
            from langchain_openai import ChatOpenAI
        except ImportError as err:
            raise ImportError(
                "Please install the required packages to run this function."
            ) from err

        browser = Browser(config=BrowserConfig(headless=True, disable_security=True))

        agent = Agent(
            task=f"""
            1. Go to Google.com.
            2. Search {self.author_name} email address in the {self.keyword_index} field and enter.
            3. Output only the emails or 'None'—no additional explanations.
            """,
            llm=ChatOpenAI(model="gpt-4o"),
            browser=browser,
        )
        loop = asyncio.get_event_loop()
        result = loop.run_until_complete(agent.run())
        return result.final_result().split("\n")


class FindSimilarity:
    """Find similarity between the author names and emails."""

    def preprocess_emails(self, email_list: List) -> Optional[str]:
        """Preprocess the email address to extract email author name."""
        return " ".join([email.split("@")[0].lower() for email in email_list])

    def find_email_author_and_save(self, list_of_emails: List[str]) -> object:
        """Find the email address and author name using Cosine Similarity."""
        print("\n📍 Step 9: Finding Similarity and Saving to Excel!")
        wb, ws = _load_excel("arxiv_scraped_data_backup.xlsx")
        headers = [
            str(cell.value).strip().lower() if cell.value else None for cell in ws[1]
        ]

        if "email" in headers:
            email_column = headers.index("email") + 1
        else:
            email_column = ws.max_column + 1
            ws.cell(row=1, column=email_column, value="Email")

        for email in list_of_emails:
            if (
                email.__contains__("None")
                or email.startswith("protected")
                or email.__contains__("*")
            ):
                continue

            print(f"Processing email: {email}")
            doc1 = self.preprocess_emails([email])
            match_found = False

            for row_index, row in enumerate(
                ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2
            ):
                author_name = row[0]
                if not author_name:
                    continue

                doc2 = str(author_name)

                vectorizer = TfidfVectorizer(
                    analyzer="char", ngram_range=(1, 2), lowercase=True
                )
                tfidf_matrix = vectorizer.fit_transform([doc1, doc2])

                cosine_sim = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])

                if cosine_sim[0][0] > 0.6:
                    print(f"Match Found: {author_name} | {cosine_sim[0][0]}")
                    current_email = ws.cell(row=row_index, column=email_column).value
                    if current_email:
                        new_email = f"{current_email}, {email}"
                        ws.cell(row=row_index, column=email_column, value=new_email)
                    else:
                        ws.cell(row=row_index, column=email_column, value=email)
                    match_found = True

            if not match_found:
                last_row = ws.max_row + 1
                ws.cell(row=last_row, column=email_column, value=email)

        return wb.save("arxiv_scraped_data_backup.xlsx")


def extract():
    # Step 1: Download PDFs
    pdf_downloader = DownloadPDF("arxiv_scraped_data.xlsx")
    pdf_downloader.start_download()

    # # Step 2: Load PDFs and extract text
    # pdf_loader = LoadPDF()
    # pdf_length = pdf_loader.__len__()
    # pdf_texts = pdf_loader.extract_text_from_pdf()

    # for i in range(pdf_length):
    #     print(f"     [INFO] Extracted text from PDF {i+1}")

    #     # Step 3: Extract email addresses from the text
    #     email_extractor = ExtractEmails()
    #     list_of_emails = email_extractor.chatgpt_response(pdf_texts)

    #     # Step 4: Find similarity between the author names and emails
    #     similarity_finder = FindSimilarity()
    #     similarity_finder.find_email_author_and_save(list_of_emails)
    #     print("-----" * 15)


def fill_empty_emails_with_search():
    wb, ws = _load_excel("arxiv_scraped_data_backup.xlsx")

    headers = [cell.value for cell in ws[1]]
    if "emails" not in headers:
        ws.cell(row=1, column=7, value="emails")

    # WebSearch using Perplexity
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        author_name = row[0]
        email_row = row[6] if len(row) > 6 else None
        if email_row is not None:
            print(
                f"     [INFO] Email already exists for this [Author: {author_name}]. Skipping..."
            )
            continue

        # Step 5: Search for email addresses using the browser-use
        web_search = WebSearch(name=str(author_name))
        list_of_emails = web_search.perplexity_search()

        # Step 6: Find similarity between the author names and emails
        similarity_finder = FindSimilarity()
        similarity_finder.find_email_author_and_save(list_of_emails)
        print("-----" * 15)

    ask_msg = input(
        "Do you want to fill empty emails row again with browser-use? [y/n]: "
    )
    if ask_msg.lower() == "y":
        # WebSearch using Browser-Use
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            author_name = row[0]
            email_row = row[6] if len(row) > 6 else None
            if email_row is not None:
                print(
                    f"     [INFO] Email already exists for this [Author: {author_name}]. Skipping..."
                )
                continue

            # Step 5: Search for email addresses using the browser-use
            web_search = WebSearch(name=str(author_name))
            list_of_emails = web_search.browser_use()

            # Step 6: Find similarity between the author names and emails
            similarity_finder = FindSimilarity()
            similarity_finder.find_email_author_and_save(list_of_emails)
            print("-----" * 15)
    else:
        print("     [INFO] Skipping browser-use search for email addresses.")


if __name__ == "__main__":
    extract()
    # fill_empty_emails_with_search()
