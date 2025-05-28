import asyncio
import concurrent.futures
import os
import subprocess
from typing import Dict, List, Optional

import aiohttp
import requests
from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright


async def _load_excel(filename: str):
    wb = load_workbook(filename=filename)
    ws = wb.active
    return ws


def _save_excel(author_list: List[Dict], link: str) -> Optional[str]:
    """Save data to Excel and return the path to the saved file"""
    if not author_list:
        print("     [INFO] No author details found!")
        return None

    # Define output paths
    output_dir = "/root/AI-Agent-and-Scraping-Automation/arxiv_project/output"
    output_file = os.path.join(output_dir, "arxiv_results.xlsx")

    try:
        # Ensure the output directory exists
        os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "Author",
                "AuthorID",
                "PaperID",
                "LastPaperLink",
                "AmountOfMentions",
                "Keyword",
            ],
        )

        for author in author_list:
            ws.append(
                [
                    author.get("Author", "null"),
                    author.get("AuthorID", "null"),
                    author.get("PaperID", "null"),
                    author.get("LastPaperLink", "null"),
                    author.get("AmountOfMentions", "null"),
                    link,
                ],
            )

        wb.save(output_file)
        print("    ✅ [INFO] Details successfully saved to Excel!")
        return output_file
    except Exception as e:
        print(f"    ❌ [ERROR] An error occurred while saving Excel file: {e}")
        return None


class ArxivScraper:
    def __init__(self, link: str):
        self.page = None
        self.link = link
        self.papers: List[Dict] = []

    async def pagination(self) -> bool:
        print("     [INFO] Checking for next page...")
        next_page_locator = self.page.locator(
            'xpath=//*[@id="main-container"]/div[2]/nav[1]/a[2]',
        )
        try:
            await next_page_locator.wait_for(state="visible", timeout=10000)
            next_class = await next_page_locator.get_attribute("class")
            if next_class and "disabled" in next_class:
                print("    ✅ [INFO] No more pages.")
                return False
            await next_page_locator.click()
            await asyncio.sleep(1)
            return True
        except Exception:
            print("     [INFO] End of the page.")
            return False

    async def connect_to_arxiv_and_search(self):
        print("\n📍 Step 1: Connecting to Arxiv!")

        try:
            await self.page.goto(self.link, wait_until="load")
            await asyncio.sleep(2)
            print("     ✅ [INFO] Arxiv connection successful!")
        except Exception as e:
            print(f"     ❌ [ERROR] An error occurred during navigation: {e}")

    # NOTE: The cosine_similarity method ideal candidate to scrap paper tags and
    # inspect to if paper content is related with tag

    # async def cosine_similarity(self, references: str) -> bool:
    #     print("DEBUG: Using cosine similarity to find related papers...")
    #     match_found = False
    #     try:
    #         from sklearn.feature_extraction.text import TfidfVectorizer
    #         from sklearn.metrics.pairwise import cosine_similarity
    #     except ImportError:
    #         print("     ❌ [ERROR] Please install scikit-learn to use this feature.")
    #         return False
    #     vectorizer = TfidfVectorizer(analyzer='char', ngram_range=(1, 2), lowercase=True)
    #     tfidf_matrix = vectorizer.fit_transform([references, " ".join(self.keyword)])

    #     cosine_sim = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])
    #     print(f'     [INFO] References: {references} | {" ".join(self.keyword)}')
    #     if cosine_sim[0][0] > 0.4:
    #         match_found = True
    #     if match_found:
    #         print(f"     [INFO] Cosine similarity: {cosine_sim[0][0]}")
    #         return True
    #     else:
    #         print(f"     [INFO] Cosine similarity: {cosine_sim[0][0]}")
    #         return False

    async def get_paper_ids(
        self,
        max_pages_DEBUG_MODE: Optional[int] = None,
    ) -> List[Dict]:
        print("\n📍 Step 2: Scraping paper links!")
        page_count = 0
        if max_pages_DEBUG_MODE:
            print(
                f"     ⚠️ [INFO] Debug mode enabled. Scraping only {max_pages_DEBUG_MODE} pages.",
            )

        while True:
            page_count += 1
            if max_pages_DEBUG_MODE and page_count > max_pages_DEBUG_MODE:
                break

            try:
                paper_entries = self.page.locator(
                    'xpath=//*[@id="main-container"]/div[2]/ol/li',
                )
            except Exception as e:
                print(f"     ❌ [ERROR] Waiting for paper entries failed: {e}")
                break

            papers = await paper_entries.all()
            if not papers:
                print("     [INFO] No paper entries found on this page.")
                break

            for paper in papers:
                try:
                    paper_link = await paper.locator(
                        'xpath=//p[@class="list-title is-inline-block"]/a',
                    ).get_attribute("href")
                    author_locator = paper.locator(
                        'xpath=//p[contains(@class, "authors")]',
                    )
                    author_text = await author_locator.text_content()

                    # NOTE: This is a placeholder for the cosine similarity check
                    # paper_link_page = await self.page.context.new_page()
                    # await paper_link_page.goto(paper_link, wait_until='load')
                    # related_keyword = paper_link_page.locator('xpath=//*[@id="abs"]/div[4]/table/tbody/tr[1]/td[2]/span')
                    # related_keyword_text = await related_keyword.text_content()
                    # await paper_link_page.close()

                    # if not await self.cosine_similarity(related_keyword_text): continue

                    paper_id = None
                    if paper_link:
                        paper_id = paper_link.split("/")[-1]

                    if not paper_id:
                        print(
                            "     ⚠️ [WARNING] Paper ID missing for one entry, skipping...",
                        )
                        continue

                    if author_text:
                        author_names = author_text.split(":")[1].split(",")
                        for author in author_names:
                            self.papers.append(
                                {
                                    "PaperID": paper_id,
                                    "Author": author.strip(),
                                    "AuthorID": 0,
                                    "LastPaperLink": "",
                                    "AmountOfMentions": 0,
                                    "Link": self.link,
                                },
                            )

                except Exception as e:
                    print(f"     ❌ [ERROR] Failed to process a paper entry: {e}")

            print(f"     [INFO] Scraped page {page_count}")
            if not await self.pagination():
                break

        print("     ✅ [INFO] Paper links scraped successfully!")
        return self.papers

    async def get_author_details(self) -> Optional[List[Dict]]:
        print("\n📍 Step 3: Scraping author details!")
        if not self.papers:
            print("     [INFO] No paper links found!")
            return None
        async with aiohttp.ClientSession() as session:
            for paper in self.papers:
                await asyncio.sleep(2)
                api = (
                    f"https://api.semanticscholar.org/v1/paper/arXiv:{paper['PaperID']}"
                    f"?include_unknown_references=true"
                )
                try:
                    async with session.get(api) as response:
                        response.raise_for_status()
                        data = await response.json()

                        if "authors" not in data:
                            print(
                                f"    [WARNING] No authors found for paper {paper['PaperID']}",
                            )
                            continue

                        for author in data["authors"]:
                            if (
                                author.get("name", "").split(" ")[-1]
                                == paper["Author"].split(" ")[-1]
                            ):
                                paper["AuthorID"] = author.get("authorId", "null")

                        print(
                            f"     [INFO] Processed authors for paper {paper['PaperID']}",
                        )
                except aiohttp.ClientResponseError as e:
                    print(
                        f"    ❌ [ERROR] HTTP error for paper {paper['PaperID']}: {e}",
                    )
                except KeyError as e:
                    print(
                        f"    ❌ [ERROR] Missing key in response for paper {paper['PaperID']}: {e}",
                    )
                except Exception as e:
                    print(
                        f"    ❌ [ERROR] An error occurred for paper {paper['PaperID']}: {e}",
                    )
                await asyncio.sleep(1)

        if not any(paper["AuthorID"] for paper in self.papers):
            print("     [INFO] No author details found for any papers.")
            return None

        print("    ✅ [INFO] Author details scraped successfully!")
        return self.papers

    async def get_amount_of_mentions(self) -> Optional[List[Dict]]:
        print("\n📍 Step 4: Using Scholar API to extract amount of mentions!")
        if not self.papers:
            print("     [INFO] No author details found!")
            return None

        async with aiohttp.ClientSession() as session:
            for row_info in self.papers:
                author_id = row_info.get("AuthorID")
                if author_id == "null":
                    print(
                        f"    [INFO] No valid author name for {row_info['Author']}. Skipping.",
                    )
                    continue

                # Create the API URL with the first author's authorId (or pick another strategy to query)
                api = (
                    f"https://api.semanticscholar.org/graph/v1/author/{author_id}"
                    f"/papers?fields=url,publicationDate,authors&limit=1000"
                )
                try:
                    async with session.get(api) as response:
                        response.raise_for_status()
                        data = await response.json()

                        if "data" in data and data["data"]:
                            amount_of_mentions = data["data"]  # List of papers
                            print(
                                f"    📄 {row_info['Author']} -> {len(amount_of_mentions)} papers found.",
                            )

                            # Update the amount of mentions for the author
                            row_info["AmountOfMentions"] = len(amount_of_mentions)
                        else:
                            print(
                                f"    ⚠️ [WARNING] No papers found for author {row_info['Author']}",
                            )

                        await asyncio.sleep(2)  # Sleep to prevent hitting rate limits

                except aiohttp.ClientResponseError as e:
                    print(
                        f"    ❌ [ERROR] HTTP error for author {row_info['Author']}: {e}",
                    )
                except KeyError as e:
                    print(
                        f"    ❌ [ERROR] Missing key in response for author {row_info['Author']}: {e}",
                    )
                except Exception as e:
                    print(
                        f"    ❌ [ERROR] An error occurred for author {row_info['Author']}: {e}",
                    )

        print("    ✅ [INFO] Related papers scraped successfully!")
        return self.papers

    async def get_related_paper_details(self):
        print("\n📍 Step 5: Scraping related paper details!")
        if not self.papers:
            print("     [INFO] No author details found!")
            return None

        for row in self.papers:
            author = row.get("Author")
            if not author:
                print(f"     [INFO] No author name found for row: {row}")
                continue

            try:
                # Enter the author's name in the search box
                search_box = self.page.locator(
                    "xpath=/html/body/header/div[2]/div[2]/form/div/div[1]/input",
                )
                await search_box.fill(author)
                await search_box.press("Enter")
                await asyncio.sleep(2)  # Wait for the search results to load

                # Locate the latest paper
                paper_entries = self.page.locator(
                    'xpath=//*[@id="main-container"]/div[2]/ol/li',
                )
                papers = await paper_entries.all()
                if not papers:
                    print(f"     [INFO] No papers found for author: {author}")
                    continue

                # Get the link of the first paper (latest paper)
                try:
                    paper_link = (
                        await papers[0]
                        .locator('xpath=//p[@class="list-title is-inline-block"]/a')
                        .get_attribute("href")
                    )
                    if paper_link:
                        print(f"     🔗 Found paper link for {author}: {paper_link}")
                        row["LastPaperLink"] = paper_link
                    else:
                        print(f"     ⚠️ [WARNING] No paper link found for {author}")
                except Exception as e:
                    print(
                        f"     ❌ [ERROR] Failed to process the first paper entry for {author}: {e}",
                    )

            except Exception as e:
                print(
                    f"     ❌ [ERROR] An error occurred while processing {author}: {e}",
                )

        print("    ✅ [INFO] Last paper links scraped successfully!")
        return self.papers


class DownloadPDF:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.proxy = {
            "account_id": "hl_fc29e1f2",
            "zone_name": "semanticscholar",
            "password": "p8tgqocdlqav",
        }

    def download_arxiv_pdf(
        self,
        pdf_url: str,
        save_dir: str = "/root/AI-Agent-and-Scraping-Automation/arxiv_project/output/pdfs/",
    ) -> bool:
        """Download PDF from the given URL using requests."""
        if not pdf_url:
            print(f"    ⚠️ [WARNING] Skipping invalid paper url: {pdf_url}")
            return False

        # Convert abstract URL to PDF URL
        if "arxiv.org/abs" in pdf_url:
            pdf_url = pdf_url.replace("arxiv.org/abs", "arxiv.org/pdf") + ".pdf"

        save_path = os.path.join(save_dir, f"{pdf_url.split('/')[-1]}")
        os.makedirs(save_dir, exist_ok=True)

        if os.path.exists(save_path):
            print(f"    ⚠️ [WARNING] {pdf_url}: PDF already exists.")
            return True

        try:
            proxy_url = "brd.superproxy.io:33335"
            proxy_auth = f"brd-customer-{self.proxy['account_id']}-zone-{self.proxy['zone_name']}:{self.proxy['password']}"

            cmd = [
                "curl",
                "--proxy",
                proxy_url,
                "--proxy-user",
                proxy_auth,
                "-k",
                pdf_url,
                "-o",
                save_path,
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)
            if result.returncode != 0:
                print(
                    f"    ❌ [ERROR] {pdf_url}: PDF can't be downloaded. Error: {result.stderr}",
                )
                return False

            print(
                f"    ✅ [INFO] {pdf_url}: PDF successfully downloaded -> {save_path}",
            )
            return True

        except requests.exceptions.RequestException as e:
            print(f"    ❌ [ERROR] {pdf_url}: Request error - {str(e)}")
            return False

    async def start_download(self, max_workers: int = 3) -> None:
        """Start downloading PDFs with rate limiting consideration."""
        print("\n📍 Step 6: Downloading PDFs!")
        ws = await _load_excel(
            filename=self.excel_path,
        )

        if not ws:
            print("     ❌ [ERROR] Excel file not found!")
            return

        pdf_urls = [row[3].value for row in ws.iter_rows(min_row=2, max_col=4)]
        print(f"    ℹ️ [INFO] Found {len(pdf_urls)} URLs to download.")

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(self.download_arxiv_pdf, pdf_url): pdf_url
                for pdf_url in pdf_urls
            }
            for future in concurrent.futures.as_completed(futures):
                pdf_url = futures[future]
                try:
                    future.result()
                except Exception as e:
                    print(f"    ❌ [ERROR] {pdf_url}: Thread error - {e}")


async def async_main_run(link: str, proxies: Optional[Dict[str, str]] = None):
    async with async_playwright() as pw:
        args = [
            "--no-sandbox",
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--disable-background-timer-throttling",
            "--disable-popup-blocking",
            "--disable-backgrounding-occluded-windows",
            "--disable-renderer-backgrounding",
            "--disable-window-activation",
            "--disable-focus-on-load",
            "--no-first-run",
            "--no-default-browser-check",
            "--no-startup-window",
            "--window-position=0,0",
            # '--window-size=1280,1000',
        ]

        browser = await pw.chromium.launch(headless=True, args=args, proxy=proxies)
        context = await browser.new_context(viewport={"width": 1280, "height": 720})
        page = await context.new_page()

        # Create the ArxivScraper instance and assign the open page
        arxiv_scraper = ArxivScraper(
            link=link,
        )
        arxiv_scraper.page = page

        # Step 1: Connect and search
        await arxiv_scraper.connect_to_arxiv_and_search()

        # Step 2: Get paper IDs (use max_pages_DEBUG_MODE=1 for testing)
        await arxiv_scraper.get_paper_ids(max_pages_DEBUG_MODE=1)

        # Step 3: Get author details
        await arxiv_scraper.get_author_details()

        # Step 4: Get amount of mentions
        await arxiv_scraper.get_amount_of_mentions()

        # Step 5: Get related paper details
        await arxiv_scraper.get_related_paper_details()

        # step 6: Save the scraped details to an Excel file
        # excel_path = _save_excel(arxiv_scraper.papers, arxiv_scraper.link)

        # # Step 7: Download PDFs
        # if excel_path:
        #     download_pdf = DownloadPDF(
        #         excel_path=excel_path,
        #     )
        #     await download_pdf.start_download()
        # else:
        #     print("     ❌ [ERROR] Excel file not created, skipping PDF download.")


if __name__ == "__main__":
    asyncio.run(
        async_main_run(
            link="https://arxiv.org/search/?query=machine+learning+AND+biology&searchtype=all&source=header",
        ),
    )
