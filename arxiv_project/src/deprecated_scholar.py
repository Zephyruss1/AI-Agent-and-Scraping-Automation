import asyncio
import os
from typing import Optional
from urllib.request import urlretrieve

from openpyxl import load_workbook
from playwright.async_api import async_playwright


class SemanticScholarScraper:
    def __init__(self, proxy_config: Optional[dict] = None):
        self.proxy_config = proxy_config or {
            "server": "brd.superproxy.io:33335",
            "username": "brd-customer-hl_fc29e1f2-zone-semanticscholar",
            "password": "p8tgqocdlqav",
        }

    @staticmethod
    def check_postfix(url: str) -> str:
        return url if url.endswith(".pdf") else url + ".pdf"

    @staticmethod
    def get_filename_from_url(url: str) -> str:
        return url.split("/")[-1]

    async def click_dropdown_control(self, page) -> bool:
        try:
            # Wait for and check if the dropdown control div is visible
            dropdown_control = page.locator(
                "div.dropdown-filters__sort-control.search-sort.flex-row-vcenter",
            )

            try:
                await dropdown_control.wait_for(state="visible", timeout=5000)
                is_visible = await dropdown_control.is_visible()
                print(f"Dropdown control visibility: {is_visible}")

                if is_visible:
                    # Click the dropdown control
                    await dropdown_control.click()
                    print("Successfully clicked dropdown control")

                    # Wait for select element to be visible after clicking
                    select_element = page.locator(
                        'xpath=//*[@id="app"]/div[1]/div[2]/div/main/div[2]/div/div/div/div[1]/div/div/div[2]/div[1]/select',
                    )
                    await select_element.wait_for(state="visible", timeout=5000)
                    return True
                else:
                    print("Dropdown control is not visible")
                    return False

            except Exception as e:
                print(f"Error while waiting for dropdown visibility: {e}")
                return False

        except Exception as e:
            print(f"Error in click_dropdown_control: {e}")
            return False

    async def select_sort_option(self, page) -> bool:
        try:
            # Use select_option on the select element
            select_element = page.locator(
                'xpath=//*[@id="app"]/div[1]/div[2]/div/main/div[2]/div/div/div/div[1]/div/div/div[2]/div[1]/select',
            )
            await select_element.select_option("pub-date")

            # Wait for network idle after selection
            await page.wait_for_load_state("networkidle")

            # Wait a bit longer for any client-side updates
            await page.wait_for_timeout(3000)

            # Check if the URL contains the sort parameter
            current_url = page.url
            if "sort=pub-date" not in current_url:
                # If not in URL, try to modify it
                new_url = (
                    current_url + ("&" if "?" in current_url else "?") + "sort=pub-date"
                )
                await page.goto(new_url, wait_until="networkidle")
                await page.wait_for_timeout(2000)

            # Verify that the sorting option is still selected
            selected_value = await select_element.evaluate("element => element.value")
            return selected_value == "pub-date"
        except Exception as e:
            print(f"Error selecting sort option: {e}")
            return False

    async def wait_for_sort_selector(self, page) -> bool:
        try:
            # First click the dropdown control
            if not await self.click_dropdown_control(page):
                print("Failed to click dropdown control")
                return False

            # Try to select the sort option
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    if await self.select_sort_option(page):
                        print("Successfully sorted by publication date")

                        # Wait for the page to settle
                        await page.wait_for_load_state("networkidle")
                        await page.wait_for_timeout(2000)

                        # Verify sorting by checking dates of first few papers
                        paper_dates = await page.evaluate("""() => {
                            const papers = document.querySelectorAll('[data-paper-id]');
                            return Array.from(papers).slice(0, 5).map(paper => {
                                const dateElement = paper.querySelector('span[data-selenium-selector="paper-year"]');
                                return dateElement ? dateElement.textContent : null;
                            });
                        }""")

                        if paper_dates and all(paper_dates):
                            # Convert to integers and check if they're in descending order
                            dates = [
                                int(date)
                                for date in paper_dates
                                if date and date.isdigit()
                            ]
                            if dates == sorted(dates, reverse=True):
                                print("Verified papers are sorted by date")
                                return True

                    print(
                        f"Attempt {attempt + 1}: Sorting verification failed, retrying...",
                    )
                    await asyncio.sleep(2)

                except Exception as e:
                    print(f"Attempt {attempt + 1} failed: {e}")
                    if attempt < max_retries - 1:
                        await asyncio.sleep(2)

            print("Failed to sort by publication date after all attempts")
            return False

        except Exception as e:
            print(f"Error in wait_for_sort_selector: {e}")
            return False

    async def download_paper(self, href: str) -> bool:
        if not href:
            print("Warning: href is None, skipping download.")
            return False

        try:
            filename = self.get_filename_from_url(href)
            print(f"Downloading file from {href}...")
            urlretrieve(href, os.path.join(os.getcwd(), filename))
            print(f"File downloaded and saved as {filename}")
            return True
        except Exception as e:
            print(f"Download error: {e}")
            return False

    async def process_paper_links(self, page) -> bool:
        link_selectors = {
            "arxiv": {"selector": 'a[href*="arxiv.org/pdf"]', "name": "arxiv link"},
            "open_pdf": {
                "selector": 'div.paper-badge-list a[href*="arxiv.org/pdf"]',
                "name": "clickable paper",
            },
            "acl": {
                "selector": ".class-cl-paper-action__button-container",
                "name": "acl link",
            },
            "close_pdf": {
                "selector": 'div.flex-row.paper-badge-list a[href*="arxiv.org/pdf"]',
                "name": "close pdf paper",
            },
            "ieee": {
                "selector": '//div[contains(@class, "paper-badge-list")]//a[contains(@href, "ieee")]',
                "name": "ieee link",
                "is_xpath": True,
            },
        }

        for _link_type, link_info in link_selectors.items():
            try:
                if link_info.get("is_xpath"):
                    locator = page.locator(f"xpath={link_info['selector']}").first
                else:
                    locator = page.locator(link_info["selector"]).first

                if await locator.count() > 0:
                    print(f"Processing with {link_info['name']}")
                    href = await locator.get_attribute("href")
                    if await self.download_paper(href):
                        return True
            except Exception as e:
                print(f"Error processing {link_info['name']}: {e}")
                continue

        return False

    async def scrape_papers(self, id_list: list[str]):
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(
                headless=False,
                proxy=self.proxy_config,
            )

            context = await browser.new_context(
                viewport={"width": 1280, "height": 720},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            )

            try:
                for author_id in id_list:
                    print(f"Processing author ID: {author_id}")
                    page = await context.new_page()

                    try:
                        url = f"http://www.semanticscholar.org/author/{author_id}"
                        await page.goto(url, wait_until="networkidle")

                        # Wait for sort selector and try to sort
                        if not await self.wait_for_sort_selector(page):
                            print("Failed to sort papers by publication date")
                            continue

                        if not await self.select_sort_option(page):
                            print("Failed to sort papers by publication date")
                            continue

                        # Process paper links
                        if not await self.process_paper_links(page):
                            print(
                                f"No papers found or downloaded for author {author_id}",
                            )

                    except Exception as e:
                        print(f"Error processing author {author_id}: {e}")
                    finally:
                        await page.close()

            except Exception as e:
                print(f"Fatal error: {e}")
            finally:
                await context.close()
                await browser.close()


async def _load_excel():
    wb = load_workbook(filename="arxiv_scraped_data.xlsx")
    ws = wb.active
    return ws


async def main():
    scraper = SemanticScholarScraper()
    ws = await _load_excel()
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        author_id = row[1]
        if author_id:
            await scraper.scrape_papers([author_id])
        print("-----------" * 15)


if __name__ == "__main__":
    asyncio.run(main())
